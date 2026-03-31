from pathlib import Path
from typing import List, Dict
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


INPUT_SHEET = "input"
RESULT_SHEET = "result"

FIO_ALIASES = {"fio", "фио"}
IIN_ALIASES = {"iin", "иин"}


class ExcelValidationError(Exception):
    pass


def _normalize(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_header(value) -> str:
    return _normalize(value).lower()


def _find_required_columns(headers: List[str]) -> Dict[str, int]:
    fio_idx = None
    iin_idx = None

    for idx, header in enumerate(headers, start=1):
        if header in FIO_ALIASES and fio_idx is None:
            fio_idx = idx
        if header in IIN_ALIASES and iin_idx is None:
            iin_idx = idx

    if fio_idx is None:
        raise ExcelValidationError(
            "Не найден обязательный столбец ФИО. Допустимые названия: fio / фио"
        )

    if iin_idx is None:
        raise ExcelValidationError(
            "Не найден обязательный столбец ИИН. Допустимые названия: iin / иин"
        )

    return {
        "fio": fio_idx,
        "iin": iin_idx,
    }


def read_people(file_path: Path) -> List[Dict]:
    wb = load_workbook(file_path)

    if INPUT_SHEET not in wb.sheetnames:
        raise ExcelValidationError(
            f"Лист '{INPUT_SHEET}' не найден. Бот читает только лист '{INPUT_SHEET}'."
        )

    ws = wb[INPUT_SHEET]
    headers = [_normalize_header(cell.value) for cell in ws[1]]
    header_map = _find_required_columns(headers)

    people = []
    seen_iins = set()

    for row_idx in range(2, ws.max_row + 1):
        fio = _normalize(ws.cell(row_idx, header_map["fio"]).value)
        iin = _normalize(ws.cell(row_idx, header_map["iin"]).value)

        if not fio and not iin:
            continue

        if not fio:
            raise ExcelValidationError(f"Пустое ФИО в строке {row_idx}")

        if not iin:
            raise ExcelValidationError(f"Пустой ИИН в строке {row_idx}")

        iin = iin.replace(" ", "").replace("\u00A0", "")

        if not iin.isdigit() or len(iin) != 12:
            raise ExcelValidationError(f"Некорректный ИИН в строке {row_idx}: {iin}")

        if iin in seen_iins:
            raise ExcelValidationError(f"Дубликат ИИН в строке {row_idx}: {iin}")

        seen_iins.add(iin)

        people.append({
            "row_number": row_idx,
            "fio": fio,
            "iin": iin,
        })

    if not people:
        raise ExcelValidationError("На листе 'input' нет строк для обработки")

    return people


def _format_worksheet(ws) -> None:
    header_font = Font(bold=True)
    header_alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )
    data_alignment = Alignment(
        horizontal="left",
        vertical="top",
        wrap_text=True
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_alignment

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = data_alignment

    for column_cells in ws.columns:
        max_length = 0
        column_index = column_cells[0].column
        column_letter = get_column_letter(column_index)

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            lines = value.splitlines() if value else [""]
            longest_line = max((len(line) for line in lines), default=0)
            if longest_line > max_length:
                max_length = longest_line

        adjusted_width = min(max(max_length + 2, 12), 45)
        ws.column_dimensions[column_letter].width = adjusted_width

    for row in ws.iter_rows():
        row_index = row[0].row
        max_lines = 1

        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            line_count = max(1, len(value.splitlines()))
            if line_count > max_lines:
                max_lines = line_count

        if row_index == 1:
            ws.row_dimensions[row_index].height = 24
        else:
            ws.row_dimensions[row_index].height = max(18, min(15 * max_lines, 60))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_results(source_file: Path, output_file: Path, results: List[Dict]) -> None:
    wb = load_workbook(source_file)

    # Старый result всегда удаляем и создаем заново.
    if RESULT_SHEET in wb.sheetnames:
        del wb[RESULT_SHEET]

    result_ws = wb.create_sheet(RESULT_SHEET)

    result_ws.append([
        "ФИО",
        "ИИН",
        "Статус обработки",
        "Выезд запрещен",
        "У кого задолженность",
        "Контакт исполнителя",
        "Дата возбуждения",
        "Сумма долга",
        "Общая сумма",
        "Количество задолженностей",
        "Текст ошибки",
    ])

    for result in results:
        fio = result.get("fio", "")
        iin = result.get("iin", "")
        check_status = result.get("check_status", "")
        travel_status = result.get("travel_status", "")
        total_amount = result.get("total_amount", "")
        debts_count = result.get("debts_count", 0)
        error_message = result.get("error_message", "")
        details = result.get("details", [])

        if not details:
            result_ws.append([
                fio,
                iin,
                check_status,
                travel_status,
                "-",
                "-",
                "-",
                "-",
                total_amount,
                debts_count,
                error_message,
            ])
            continue

        for detail in details:
            result_ws.append([
                fio,
                iin,
                check_status,
                travel_status,
                detail.get("issuer", "-"),
                detail.get("executor_contact", "-"),
                detail.get("start_date", "-"),
                detail.get("amount", "-"),
                total_amount,
                debts_count,
                error_message,
            ])

    for ws in wb.worksheets:
        _format_worksheet(ws)

    wb.save(output_file)