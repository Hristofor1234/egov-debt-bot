import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from excel_utils import ExcelValidationError, read_people, write_results, write_single_result


class ExcelUtilsTests(unittest.TestCase):
    def make_workbook(self, rows, headers=("ФИО", "ИИН")) -> Path:
        path = Path(tempfile.mkstemp(suffix=".xlsx")[1])
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "input"
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        workbook.save(path)
        return path

    def test_reads_russian_headers_and_normalizes_iin(self):
        path = self.make_workbook([("Алия Тест", "123 456 789 012")])
        self.addCleanup(path.unlink)

        self.assertEqual(
            read_people(path),
            [{"row_number": 2, "fio": "Алия Тест", "iin": "123456789012"}],
        )

    def test_rejects_duplicate_iin(self):
        path = self.make_workbook([
            ("Алия Тест", "123456789012"),
            ("Бек Тест", "123456789012"),
        ])
        self.addCleanup(path.unlink)

        with self.assertRaisesRegex(ExcelValidationError, "Дубликат ИИН"):
            read_people(path)

    def test_writes_one_result_row_per_debt(self):
        source = self.make_workbook([("Алия Тест", "123456789012")])
        output = Path(tempfile.mkstemp(suffix=".xlsx")[1])
        self.addCleanup(source.unlink)
        self.addCleanup(output.unlink)

        write_results(source, output, [{
            "fio": "Алия Тест",
            "iin": "123456789012",
            "check_status": "Обработано",
            "travel_status": "Разрешен",
            "total_amount": "300",
            "debts_count": 2,
            "error_message": "",
            "details": [
                {"issuer": "Орган 1", "amount": "100"},
                {"issuer": "Орган 2", "amount": "200"},
            ],
        }])

        result = load_workbook(output)["result"]
        self.assertEqual(result.max_row, 3)
        self.assertEqual(result["E2"].value, "Орган 1")
        self.assertEqual(result["E3"].value, "Орган 2")
        self.assertEqual(result["I2"].value, "300")

    def test_writes_standalone_chat_result(self):
        output = Path(tempfile.mkstemp(suffix=".xlsx")[1])
        self.addCleanup(output.unlink)

        write_single_result(output, {
            "fio": "Алия Тест",
            "iin": "123456789012",
            "check_status": "Не найдено",
            "travel_status": "-",
            "total_amount": "-",
            "debts_count": 0,
            "error_message": "",
            "details": [],
        })

        workbook = load_workbook(output)
        self.assertEqual(workbook.sheetnames, ["input", "result"])
        self.assertEqual(workbook["result"]["C2"].value, "Не найдено")


if __name__ == "__main__":
    unittest.main()
